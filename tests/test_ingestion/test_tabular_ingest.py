"""Integration tests for the spreadsheet structured-ingest orchestrator."""
import json
from unittest.mock import MagicMock

import openpyxl
import pytest

from src.db.metadata import MetadataStore
from src.db.schema_registry import SchemaRegistry
from src.ingestion.tabular_store import connect_tabular, duckdb_table_name


def _write_xlsx(path, sheets):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    wb.save(path)


def _fake_profile_generate(system_prompt, user_prompt, **kwargs):
    return json.dumps({
        "column_descriptions": {"grade": "Pay grade", "step": "Step", "salary": "Annual salary"},
        "key_columns": ["grade", "step"],
        "measure_columns": ["salary"],
        "table_description": "GS pay by grade and step",
    })


@pytest.mark.asyncio
async def test_ingest_clean_sheet_stores_rows_schema_and_narratives(tmp_path, monkeypatch):
    from src.config import settings
    monkeypatch.setattr(settings, "tabular_duckdb_path", str(tmp_path / "t.duckdb"))

    # embed_texts is patched where the orchestrator looks it up
    import src.ingestion.tabular_ingest as ti
    monkeypatch.setattr(ti, "embed_texts", lambda texts: [[0.1, 0.2, 0.3] for _ in texts])

    xlsx = tmp_path / "pay.xlsx"
    _write_xlsx(xlsx, {
        "Pay": [["grade", "step", "salary"]] + [[f"GS-{g}", 5, 80000 + g] for g in range(10, 14)],
    })

    vector_store = MagicMock()
    store = MetadataStore(database_url=f"sqlite+aiosqlite:///{tmp_path}/m.db")
    await store.init()
    registry = SchemaRegistry()

    from src.ingestion.tabular_ingest import ingest_spreadsheet_tables
    n = await ingest_spreadsheet_tables(
        str(xlsx), "doc1", "pay.xlsx", "xlsx", ["ALL"], "",
        vector_store, store, schema_registry=registry, generate_fn=_fake_profile_generate,
    )

    assert n == 1
    # narratives embedded + upserted, one per data row, marked as table_row
    vector_store.upsert.assert_called_once()
    kwargs = vector_store.upsert.call_args.kwargs
    assert len(kwargs["texts"]) == 4
    assert all("Pay grade=" in t for t in kwargs["texts"])
    assert all(m.chunk_size_tier == "table_row" for m in kwargs["metadatas"])
    # schema registered in-memory AND persisted
    assert registry.list_for_user(["ALL"])
    persisted = await store.load_all_schemas()
    assert persisted[0].table == duckdb_table_name("doc1", "Pay")
    # rows landed in DuckDB
    con = connect_tabular(read_only=True)
    cnt = con.execute(f'SELECT COUNT(*) FROM "{duckdb_table_name("doc1", "Pay")}"').fetchone()[0]
    con.close()
    assert cnt == 4


@pytest.mark.asyncio
async def test_messy_only_workbook_stores_nothing(tmp_path, monkeypatch):
    from src.config import settings
    monkeypatch.setattr(settings, "tabular_duckdb_path", str(tmp_path / "t.duckdb"))
    import src.ingestion.tabular_ingest as ti
    monkeypatch.setattr(ti, "embed_texts", lambda texts: [[0.1] for _ in texts])

    xlsx = tmp_path / "notes.xlsx"
    _write_xlsx(xlsx, {"Readme": [["This is a narrative note."], ["Updated 2024."]]})

    vector_store = MagicMock()
    store = MetadataStore(database_url=f"sqlite+aiosqlite:///{tmp_path}/m.db")
    await store.init()
    registry = SchemaRegistry()

    from src.ingestion.tabular_ingest import ingest_spreadsheet_tables
    n = await ingest_spreadsheet_tables(
        str(xlsx), "doc2", "notes.xlsx", "xlsx", ["ALL"], "",
        vector_store, store, schema_registry=registry, generate_fn=_fake_profile_generate,
    )
    assert n == 0
    vector_store.upsert.assert_not_called()
    assert await store.load_all_schemas() == []


@pytest.mark.asyncio
async def test_per_sheet_failure_is_contained(tmp_path, monkeypatch):
    from src.config import settings
    monkeypatch.setattr(settings, "tabular_duckdb_path", str(tmp_path / "t.duckdb"))
    import src.ingestion.tabular_ingest as ti
    # embedding blows up -> the clean sheet fails, but the call must not raise
    monkeypatch.setattr(ti, "embed_texts", lambda texts: (_ for _ in ()).throw(RuntimeError("embed down")))

    xlsx = tmp_path / "pay.xlsx"
    _write_xlsx(xlsx, {
        "Pay": [["grade", "step", "salary"]] + [[f"GS-{g}", 5, 80000 + g] for g in range(10, 14)],
    })

    vector_store = MagicMock()
    store = MetadataStore(database_url=f"sqlite+aiosqlite:///{tmp_path}/m.db")
    await store.init()
    registry = SchemaRegistry()

    from src.ingestion.tabular_ingest import ingest_spreadsheet_tables
    n = await ingest_spreadsheet_tables(  # must NOT raise
        str(xlsx), "doc3", "pay.xlsx", "xlsx", ["ALL"], "",
        vector_store, store, schema_registry=registry, generate_fn=_fake_profile_generate,
    )
    # The sheet wrote DuckDB rows + a schema before embedding failed; that
    # partial state is acceptable (valid for SQL, idempotent on re-ingest).
    # `n` counts only FULLY-processed sheets, so it is 0 and no narratives upsert.
    assert n == 0  # the sheet failed at the embed step, so it was not counted
    vector_store.upsert.assert_not_called()


@pytest.mark.asyncio
async def test_ingest_document_invokes_tabular_branch_for_spreadsheet(tmp_path, monkeypatch):
    import src.ingestion.pipeline as pipe
    import src.ingestion.tabular_ingest as ti
    import src.generation.llm_client as llm
    import src.knowledge.graph_rag as kg

    # Patch heavy collaborators so we exercise only the branch logic.
    monkeypatch.setattr(pipe, "embed_texts", lambda texts: [[0.1] for _ in texts])
    monkeypatch.setattr(llm, "generate", lambda **kw: "summary")

    async def _noop_insert(*a, **k):
        return None
    monkeypatch.setattr(kg, "insert_document", _noop_insert)

    calls = []

    async def _fake_orch(*args, **kwargs):
        calls.append(args)
        return [], [], set()
    # pipeline imports the orchestrator into its own namespace, so patch it there.
    monkeypatch.setattr(pipe, "ingest_structured_sheets", _fake_orch)

    store = MetadataStore(database_url=f"sqlite+aiosqlite:///{tmp_path}/m.db")
    await store.init()
    vector_store = MagicMock()

    xlsx = tmp_path / "pay.xlsx"
    _write_xlsx(xlsx, {"Pay": [["grade", "salary"], ["GS-12", 86415], ["GS-13", 102000], ["GS-14", 120000]]})
    await pipe.ingest_document(str(xlsx), ["ALL"], "tester", vector_store, store)
    assert len(calls) == 1  # orchestrator invoked for the spreadsheet

    txt = tmp_path / "note.txt"
    txt.write_text("just some text\n", encoding="utf-8")
    await pipe.ingest_document(str(txt), ["ALL"], "tester", vector_store, store)
    assert len(calls) == 1  # NOT invoked again for a non-spreadsheet


@pytest.mark.asyncio
async def test_populate_schema_registry_loads_persisted(tmp_path):
    from src.db.schema_registry import TableSchema, ColumnSchema
    from src.ingestion.tabular_ingest import populate_schema_registry

    store = MetadataStore(database_url=f"sqlite+aiosqlite:///{tmp_path}/m.db")
    await store.init()
    await store.save_schema(TableSchema(
        database="spreadsheets", table="doc_x_pay",
        columns=[ColumnSchema(name="grade", dtype="VARCHAR", description="Pay grade")],
        description="GS pay", acl_groups=["ALL"],
    ))

    registry = SchemaRegistry()
    await populate_schema_registry(store, registry)

    loaded = registry.list_for_user(["ALL"])
    assert len(loaded) == 1
    assert loaded[0].table == "doc_x_pay"


def test_schema_registry_remove():
    from src.db.schema_registry import SchemaRegistry, TableSchema, ColumnSchema
    reg = SchemaRegistry()
    s = TableSchema(database="spreadsheets", table="doc_x",
                    columns=[ColumnSchema("a", "DOUBLE", "")], description="", acl_groups=["ALL"])
    reg.register(s)
    assert reg.get_schema("spreadsheets", "doc_x") is not None
    reg.remove("spreadsheets", "doc_x")
    assert reg.get_schema("spreadsheets", "doc_x") is None
    reg.remove("spreadsheets", "doc_x")  # idempotent: no error on missing key


@pytest.mark.asyncio
async def test_cleanup_spreadsheet_tables_drops_tables_and_schemas(tmp_path, monkeypatch):
    from src.config import settings
    monkeypatch.setattr(settings, "tabular_duckdb_path", str(tmp_path / "t.duckdb"))
    from src.ingestion.tabular import SheetGrid, SheetClassification
    from src.ingestion.tabular_store import load_sheet_to_duckdb, schema_from_sheet
    from src.ingestion.tabular_ingest import cleanup_spreadsheet_tables

    grid = SheetGrid("Pay", [["grade", "salary"], ["GS-12", 86415], ["GS-13", 102000], ["GS-14", 120000]])
    cls = SheetClassification("Pay", "clean", 0, ["text", "number"], "clean table")
    con = connect_tabular(read_only=False)
    table, _ = load_sheet_to_duckdb(con, "doc1", "Pay", cls, grid)
    con.close()

    store = MetadataStore(database_url=f"sqlite+aiosqlite:///{tmp_path}/m.db")
    await store.init()
    schema = schema_from_sheet("doc1", "Pay", cls, grid, acl_groups=["ALL"])
    await store.save_schema(schema)
    reg = SchemaRegistry()
    reg.register(schema)
    assert reg.get_schema(schema.database, table) is not None  # precondition

    dropped = await cleanup_spreadsheet_tables("doc1", store, schema_registry=reg)

    assert dropped == 1
    assert await store.load_all_schemas() == []               # persisted schema gone
    assert reg.get_schema(schema.database, table) is None      # live registry cleared
    con = connect_tabular(read_only=True)
    remaining = con.execute(
        "SELECT count(*) FROM information_schema.tables WHERE table_name = ?", [table]).fetchone()[0]
    con.close()
    assert remaining == 0                                      # DuckDB table dropped


@pytest.mark.asyncio
async def test_cleanup_spreadsheet_tables_noop_for_unknown_doc(tmp_path, monkeypatch):
    from src.config import settings
    monkeypatch.setattr(settings, "tabular_duckdb_path", str(tmp_path / "t.duckdb"))
    from src.ingestion.tabular_ingest import cleanup_spreadsheet_tables
    store = MetadataStore(database_url=f"sqlite+aiosqlite:///{tmp_path}/m.db")
    await store.init()
    n = await cleanup_spreadsheet_tables("nonexistent-doc", store, schema_registry=SchemaRegistry())
    assert n == 0


from src.ingestion.tabular_ingest import ingest_structured_sheets, SPREADSHEET_DOC_TYPES


class _FakeVectorStore:
    def __init__(self):
        self.upserts = []  # list of (texts, metadatas)

    def upsert(self, texts, vectors, metadatas):
        self.upserts.append((list(texts), list(metadatas)))


class _FakeMetadataStore:
    def __init__(self):
        self.saved = []

    async def save_schema(self, schema):
        self.saved.append(schema)


class _FakeRegistry:
    def __init__(self):
        self.registered = []

    def register(self, schema):
        self.registered.append(schema)


@pytest.mark.asyncio
async def test_ingest_structured_sheets_clean_ingested_messy_gets_region_narratives(tmp_path, monkeypatch):
    import openpyxl
    p = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    clean = wb.create_sheet("Pay")
    for row in [["locality", "grade", "salary"], ["Tampa", "GS-12", 86415],
                ["Boston", "GS-12", 92000], ["Denver", "GS-13", 99000]]:
        clean.append(row)
    # Genuinely messy: a clean 3-row block embedded among total/prose rows whose
    # text in the numeric salary column makes the SHEET fail clean classification,
    # while find_table_region still carves out the clean block for narratives.
    messy = wb.create_sheet("Notes")
    for row in [["2026 Pay Notes"], ["locality", "grade", "salary"],
                ["Reno", "GS-9", 55000], ["Mesa", "GS-9", 56000], ["Ames", "GS-9", 57000],
                ["Totals", "across all", "see appendix"],
                ["Average", "N/A", "varies by locality"],
                ["Footnote", "refer to", "HR schedule B"]]:
        messy.append(row)
    wb.save(p)

    # avoid real embeddings: return a fixed-width vector per text
    monkeypatch.setattr("src.ingestion.tabular_ingest.embed_texts",
                        lambda texts, *a, **k: [[0.0, 0.0, 0.0] for _ in texts])

    vs, ms, reg = _FakeVectorStore(), _FakeMetadataStore(), _FakeRegistry()
    grids, clss, ingested = await ingest_structured_sheets(
        str(p), "doc1", "book.xlsx", "xlsx", ["g1"], "cat",
        vs, ms, schema_registry=reg, generate_fn=lambda **k: "{}",
    )
    assert [g.sheet_name for g in grids] == ["Pay", "Notes"]
    assert "Pay" in ingested              # clean sheet fully ingested
    assert "Notes" not in ingested        # messy sheet not "ingested" to DuckDB
    # messy region narratives were embedded at the table_row tier
    tiers = [m.chunk_size_tier for _, metas in vs.upserts for m in metas]
    assert "table_row" in tiers
    # narrative text restates a messy-region row
    all_texts = [t for texts, _ in vs.upserts for t in texts]
    assert any("Reno" in t for t in all_texts)


@pytest.mark.asyncio
async def test_ingest_structured_sheets_read_failure_returns_empty(monkeypatch):
    vs, ms, reg = _FakeVectorStore(), _FakeMetadataStore(), _FakeRegistry()
    grids, clss, ingested = await ingest_structured_sheets(
        "/no/such/file.xlsx", "doc1", "x.xlsx", "xlsx", ["g1"], "cat",
        vs, ms, schema_registry=reg, generate_fn=lambda **k: "{}",
    )
    assert grids == [] and clss == [] and ingested == set()


from types import SimpleNamespace


@pytest.mark.asyncio
async def test_purge_orphan_schemas_removes_only_orphans(monkeypatch):
    from src.ingestion.tabular_ingest import purge_orphan_schemas
    from src.ingestion.tabular_store import duckdb_table_name

    live_tbl = duckdb_table_name("live1", "pay")
    ghost_tbl = duckdb_table_name("ghost", "pay")

    class FakeMS:
        def __init__(self):
            self.deleted = []

        async def list_documents(self, user_groups=None):
            return [SimpleNamespace(doc_id="live1", acl_groups=["ALL"])]

        async def load_all_schemas(self):
            return [
                SimpleNamespace(database="spreadsheets", table=live_tbl),
                SimpleNamespace(database="spreadsheets", table=ghost_tbl),
            ]

        async def delete_schema(self, database, table):
            self.deleted.append((database, table))

    class FakeReg:
        def __init__(self):
            self.removed = []

        def remove(self, database, table):
            self.removed.append((database, table))

    class FakeCon:
        def __init__(self, tables):
            self.tables = tables
            self.dropped = []

        def execute(self, sql):
            if sql.startswith("SELECT table_name"):
                rows = [(t,) for t in self.tables]
                return SimpleNamespace(fetchall=lambda: rows)
            if sql.startswith("DROP TABLE"):
                self.dropped.append(sql)
            return SimpleNamespace(fetchall=lambda: [])

        def close(self):
            pass

    fake_con = FakeCon([live_tbl, ghost_tbl, "system_meta"])
    monkeypatch.setattr("src.ingestion.tabular_store.connect_tabular", lambda read_only=False: fake_con)

    ms, reg = FakeMS(), FakeReg()
    removed = await purge_orphan_schemas(ms, schema_registry=reg)

    assert removed == 1
    assert ms.deleted == [("spreadsheets", ghost_tbl)]
    assert reg.removed == [("spreadsheets", ghost_tbl)]
    assert any(ghost_tbl in d for d in fake_con.dropped)
    assert not any(live_tbl in d for d in fake_con.dropped)
    assert not any("system_meta" in d for d in fake_con.dropped)


@pytest.mark.asyncio
async def test_purge_orphan_schemas_fails_open_when_list_documents_raises(monkeypatch):
    from src.ingestion.tabular_ingest import purge_orphan_schemas

    class BoomMS:
        async def list_documents(self, user_groups=None):
            raise RuntimeError("db down")

    class FakeReg:
        def __init__(self):
            self.removed = []

        def remove(self, database, table):
            self.removed.append((database, table))

    # connect_tabular must never be reached when listing docs fails.
    def boom_connect(read_only=False):
        raise AssertionError("connect_tabular should not be called on abort")

    monkeypatch.setattr("src.ingestion.tabular_store.connect_tabular", boom_connect)

    reg = FakeReg()
    removed = await purge_orphan_schemas(BoomMS(), schema_registry=reg)
    assert removed == 0
    assert reg.removed == []


@pytest.mark.asyncio
async def test_ingest_grids_processes_provided_grids(monkeypatch):
    """ingest_grids ingests grids passed directly (no file read), so PDF-derived
    grids reuse the same clean/messy logic as Excel sheets."""
    from src.ingestion import tabular_ingest as ti
    from src.ingestion.tabular import SheetGrid
    monkeypatch.setattr("src.ingestion.tabular_ingest.embed_texts",
                        lambda texts, *a, **k: [[0.0, 0.0, 0.0] for _ in texts])
    grids = [SheetGrid("p0_table0",
                       [["grade", "over2", "over4"],
                        ["O-1", "3998.40", "5031.30"],
                        ["O-2", "4606.80", "6042.90"],
                        ["E-1", "2017.20", "2017.20"]])]
    vs, ms, reg = _FakeVectorStore(), _FakeMetadataStore(), _FakeRegistry()
    classifications, ingested = await ti.ingest_grids(
        grids, "docX", "ad.pdf", "pdf", ["executives"], "payroll_compensation",
        vs, ms, schema_registry=reg,
        generate_fn=lambda **k: '{"key_columns":["grade"],"measure_columns":["over2","over4"],'
                                '"column_descriptions":{},"table_description":"AD pay"}',
    )
    assert "p0_table0" in ingested                         # clean sheet structured
    assert classifications[0].route == "clean"
    assert any(m.chunk_size_tier == "table_row"
               for _, metas in vs.upserts for m in metas)  # narratives embedded


@pytest.mark.asyncio
async def test_ingest_grids_applies_seeded_glossary_to_narratives(monkeypatch):
    from src.ingestion import tabular_ingest as ti
    from src.ingestion.tabular import SheetGrid
    from src.db.hint_store import HintStore, SchemaHint
    monkeypatch.setattr("src.ingestion.tabular_ingest.embed_texts",
                        lambda texts, *a, **k: [[0.0, 0.0, 0.0] for _ in texts])
    hint_store = HintStore()
    hint_store.register(SchemaHint(
        scope_type="category", scope_value="payroll_compensation",
        hint_type="value_glossary", target_column="grade",
        payload={"E-*": "Enlisted Member", "O-*": "Commissioned Officer"}))
    monkeypatch.setattr("src.api.routes_ingest.get_hint_store", lambda: hint_store)

    grids = [SheetGrid("p0_table0",
                       [["grade", "over2"], ["O-1", "3998"], ["O-2", "4606"], ["E-1", "2017"]])]
    vs, ms, reg = _FakeVectorStore(), _FakeMetadataStore(), _FakeRegistry()
    await ti.ingest_grids(
        grids, "docG", "ad.pdf", "pdf", ["executives"], "payroll_compensation",
        vs, ms, schema_registry=reg,
        generate_fn=lambda **k: '{"key_columns":["grade"],"measure_columns":["over2"],'
                                '"column_descriptions":{},"table_description":"AD pay"}',
    )
    joined = "\n".join(t for texts, _ in vs.upserts for t in texts)
    assert "Enlisted Member" in joined
