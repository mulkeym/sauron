"""Structured parse + clean/messy classification for spreadsheets.

Turns a spreadsheet file into per-sheet grids (lists of cell rows) and decides
whether each sheet is a clean data table (route to structured/SQL handling) or
messy/narrative (route to text RAG). Pure functions operate on in-memory grids
so they are trivially testable; ``read_sheets`` is the only file-touching entry.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from pathlib import Path

from src.ingestion.parser import _sniff_workbook_format

MAX_HEADER_SCAN = 10
HEADER_MIN_FILLED = 0.6
HEADER_MAX_NUMERIC = 0.3
MIN_HEADER_CELLS = 2
MIN_DATA_ROWS = 3
RECTANGULAR_RATIO = 0.8
COLUMN_CONSISTENCY = 0.8


def _cell_kind(value) -> str:
    """Classify a cell as 'empty', 'number', or 'text'.

    Numbers include native int/float and numeric-looking strings with common
    formatting ($, commas, %, surrounding whitespace).
    """
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "text"  # bools are not measures here
    if isinstance(value, (int, float)):
        return "number"
    s = str(value).strip()
    if not s:
        return "empty"
    cleaned = s.replace("$", "").replace(",", "").replace("%", "").strip()
    try:
        float(cleaned)
        return "number"
    except ValueError:
        return "text"


@dataclass
class SheetGrid:
    """One sheet's raw rows, exactly as read (no header interpretation yet)."""
    sheet_name: str
    rows: list[list] = field(default_factory=list)


@dataclass
class SheetClassification:
    """Routing decision for one sheet."""
    sheet_name: str
    route: str               # "clean" or "messy"
    header_row_index: int    # -1 if no header detected
    column_dtypes: list[str] = field(default_factory=list)
    reason: str = ""


def read_sheets(path: Path) -> list[SheetGrid]:
    """Read a spreadsheet/CSV into one ``SheetGrid`` per sheet.

    Dispatches on magic bytes (not extension) so a mislabeled workbook still
    parses with the right reader. .xlsx/.xlsm via openpyxl, legacy .xls via
    xlrd, everything else as delimited text.
    """
    fmt = _sniff_workbook_format(path)
    if fmt == "ooxml":
        return _read_ooxml(path)
    if fmt == "ole":
        return _read_ole(path)
    return _read_delimited(path)


def _read_ooxml(path: Path) -> list[SheetGrid]:
    from openpyxl import load_workbook

    with open(path, "rb") as fh:
        wb = load_workbook(io.BytesIO(fh.read()), read_only=True, data_only=True)
    grids = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        grids.append(SheetGrid(sheet_name=name, rows=rows))
    wb.close()
    return grids


def _read_ole(path: Path) -> list[SheetGrid]:
    import xlrd

    book = xlrd.open_workbook(str(path))
    grids = []
    for sheet in book.sheets():
        rows = [list(sheet.row_values(r)) for r in range(sheet.nrows)]
        grids.append(SheetGrid(sheet_name=sheet.name, rows=rows))
    return grids


def _read_delimited(path: Path) -> list[SheetGrid]:
    if b"\x00" in path.read_bytes()[:8192]:
        raise ValueError(f"{path.name}: appears to be binary, not delimited text; refusing to read")
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        rows = [list(r) for r in csv.reader(f, delimiter=delimiter)]
    return [SheetGrid(sheet_name=path.stem, rows=rows)]


def detect_header_row(rows: list[list]) -> int:
    """Index of the row that looks like column headers, or -1 if none.

    A header row has >= MIN_HEADER_CELLS non-empty cells (so a single-cell title
    banner is skipped), is mostly-filled, mostly-non-numeric, and is followed by
    at least one more row (the data). Scans only the first ``MAX_HEADER_SCAN``
    rows so a giant sheet stays cheap.
    """
    for i, row in enumerate(rows[:MAX_HEADER_SCAN]):
        if i + 1 >= len(rows):
            break  # a header needs at least one data row beneath it
        kinds = [_cell_kind(c) for c in row]
        if not kinds:
            continue
        filled = [k for k in kinds if k != "empty"]
        if len(filled) < MIN_HEADER_CELLS:
            continue  # single-cell title banner or near-empty row, not a header
        filled_ratio = len(filled) / len(kinds)
        numeric_ratio = sum(1 for k in filled if k == "number") / len(filled)
        if filled_ratio >= HEADER_MIN_FILLED and numeric_ratio <= HEADER_MAX_NUMERIC:
            return i
    return -1


def infer_column_dtypes(rows: list[list], header_row_index: int) -> list[str]:
    """Dominant kind ('number'/'text'/'empty') of each column's DATA cells.

    Looks only at rows below the header. A column with no non-empty data cells
    is 'empty'; otherwise the most common non-empty kind wins (ties -> 'text').
    Column count is taken from the header row.
    """
    header = rows[header_row_index] if 0 <= header_row_index < len(rows) else []
    ncols = len(header)
    data = rows[header_row_index + 1:]
    dtypes = []
    for col in range(ncols):
        counts = {"number": 0, "text": 0}
        for row in data:
            if col >= len(row):
                continue
            kind = _cell_kind(row[col])
            if kind == "empty":
                continue
            counts[kind] += 1
        if counts["number"] == 0 and counts["text"] == 0:
            dtypes.append("empty")
        elif counts["number"] > counts["text"]:
            dtypes.append("number")
        else:
            dtypes.append("text")
    return dtypes


def classify_sheet(grid: SheetGrid) -> SheetClassification:
    """Decide whether a sheet is a clean data table or messy/narrative.

    Clean requires ALL of: a detected header, >= MIN_DATA_ROWS data rows, a
    rectangular shape (>= RECTANGULAR_RATIO of data rows match header width),
    and every non-empty column being type-consistent (>= COLUMN_CONSISTENCY of
    its non-empty cells share one kind). Anything else is messy.
    """
    rows = grid.rows
    header_idx = detect_header_row(rows)
    if header_idx < 0:
        return SheetClassification(grid.sheet_name, "messy", -1, [], "no header detected")

    header = rows[header_idx]
    ncols = len(header)
    data = rows[header_idx + 1:]
    if len(data) < MIN_DATA_ROWS:
        return SheetClassification(grid.sheet_name, "messy", header_idx, [], "too few data rows")

    matching = sum(1 for r in data if len(r) == ncols)
    if matching / len(data) < RECTANGULAR_RATIO:
        return SheetClassification(grid.sheet_name, "messy", header_idx, [], "not rectangular")

    dtypes = infer_column_dtypes(rows, header_idx)
    for col in range(ncols):
        counts = {"number": 0, "text": 0}
        for r in data:
            if col >= len(r):
                continue
            kind = _cell_kind(r[col])
            if kind != "empty":
                counts[kind] += 1
        total = counts["number"] + counts["text"]
        if total == 0:
            continue  # empty column doesn't disqualify
        dominant = max(counts["number"], counts["text"])
        if dominant / total < COLUMN_CONSISTENCY:
            return SheetClassification(grid.sheet_name, "messy", header_idx, dtypes,
                                       f"column {col} not type-consistent")

    return SheetClassification(grid.sheet_name, "clean", header_idx, dtypes, "clean table")


def analyze_spreadsheet(path: Path) -> list[SheetClassification]:
    """Read a spreadsheet file and classify every sheet."""
    return [classify_sheet(grid) for grid in read_sheets(path)]
